"""
Import corrected golden list Excel files as labeling task results.
Each customer has specific filtering rules based on buyer feedback.
"""

import json
import os
import re
import sys

import openpyxl
import requests

DRIVE_BASE = (
    "/Users/seungohnam/Library/CloudStorage/GoogleDrive-nam.s@cookiedeal.io/"
    "Shared drives/Cookiedeal/Ops/아웃바운드 서비스/진행 고객/"
)

API_URL = "http://localhost:8000"


def login():
    resp = requests.post(
        f"{API_URL}/api/labeling/auth/login",
        json={"email": "admin@cookiedeal.com", "password": "test1234"},
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def get_all_tasks(token):
    tasks = {}
    page = 1
    while True:
        resp = requests.get(
            f"{API_URL}/api/labeling/tasks",
            params={"page": page, "size": 100},
            headers={"Authorization": f"Bearer {token}"},
        )
        resp.raise_for_status()
        data = resp.json()
        for t in data["items"]:
            tasks[t["query_id"]] = t["id"]
        if page >= data["pages"]:
            break
        page += 1
    return tasks


def upload_results(token, task_id, results):
    resp = requests.post(
        f"{API_URL}/api/labeling/admin/tasks/{task_id}/results",
        json={"results": results},
        headers={"Authorization": f"Bearer {token}"},
    )
    if resp.ok:
        return True, ""
    return False, resp.text[:200]


def extract_rows(ws, header_row_idx, max_data_row=None):
    """Extract company rows with metadata. Returns list of dicts."""
    headers = []
    for cell in ws[header_row_idx]:
        h = str(cell.value).strip().replace("\n", " ") if cell.value else f"col_{cell.column}"
        headers.append(h)

    # Extract fiscal year from financial column headers (e.g., "2024년 매출액(천원)")
    fiscal_year = None
    for h in headers:
        if any(kw in h for kw in ["매출", "영업이익", "자산", "자본", "부채"]):
            m = re.search(r'(\d{4})년', h)
            if m:
                fiscal_year = m.group(1)
                break

    # Find company name column index once
    name_idx = None
    for j, h in enumerate(headers):
        if "기업명" in h or "회사명" in h:
            name_idx = j
            break
    if name_idx is None:
        name_idx = 2  # fallback

    STOP_KEYWORDS = ["제안 사유", "사업 영역", "추가 제안", "참고사항", "비고", "합계"]

    companies = []
    rank = 1
    consecutive_skips = 0
    for row_cells in ws.iter_rows(min_row=header_row_idx + 1, max_row=max_data_row or ws.max_row):
        vals = [c.value for c in row_cells]
        if not any(v is not None for v in vals[:6]):
            consecutive_skips += 1
            if consecutive_skips >= 3:
                break
            continue

        name = vals[name_idx] if name_idx < len(vals) else None
        if not name or not isinstance(name, str) or not name.strip():
            consecutive_skips += 1
            if consecutive_skips >= 3:
                break
            continue

        name = name.strip()

        # Stop at non-company data
        if any(kw in name for kw in STOP_KEYWORDS):
            break
        if len(name) > 50:  # paragraphs, not company names
            continue

        consecutive_skips = 0

        meta = {}
        fill_color = None
        if name_idx < len(list(row_cells)):
            cell = list(row_cells)[name_idx]
            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.rgb:
                fill_color = str(fill.fgColor.rgb)

        for j, h in enumerate(headers):
            if j >= len(vals) or vals[j] is None:
                continue
            v = vals[j]
            # CEO: match "대표자"/"대표이사" but NOT "대표번호"/"대표전화"
            if ("대표자" in h or "대표이사" in h) and "번호" not in h and "전화" not in h:
                val_str = str(v).strip()
                if re.match(r'^[\d\-\(\)\s,+]+$', val_str):
                    meta["phone"] = val_str  # phone number, not CEO name
                else:
                    meta["ceo"] = val_str
            elif "대표번호" in h or "대표전화" in h or "연락처" in h or ("전화" in h and "대표" not in h):
                meta["phone"] = str(v)
            elif "사업자등록" in h:
                meta["brn"] = str(v)
            elif "매출" in h and "율" not in h:
                try:
                    meta["revenue"] = int(float(v))
                except (ValueError, TypeError):
                    if isinstance(v, str) and "억" in v:
                        meta["revenue_text"] = v
            elif "영업이익" in h and "율" not in h and "률" not in h:
                try:
                    meta["operating_profit"] = int(float(v))
                except (ValueError, TypeError):
                    if isinstance(v, str) and "억" in v:
                        meta["op_text"] = v
            elif "자산" in h:
                try:
                    meta["total_assets"] = int(float(v))
                except (ValueError, TypeError):
                    pass
            elif "자본" in h and "부채" not in h:
                try:
                    meta["equity"] = int(float(v))
                except (ValueError, TypeError):
                    pass
            elif "부채" in h:
                try:
                    meta["debt"] = int(float(v))
                except (ValueError, TypeError):
                    pass
            elif "기업유형" in h or "유형" in h:
                meta["company_type"] = str(v)
            elif "홈페이지" in h:
                meta["homepage"] = str(v)
            elif "업종" in h:
                meta["industry"] = str(v)
            elif "추천 사유" in h or "추천사유" in h:
                meta["recommendation"] = str(v)
            elif "구분" in h:
                meta["category"] = str(v)
            elif "생산품" in h or "주요" in h:
                meta["products"] = str(v)
            elif "선호도" in h:
                meta["preference"] = str(v)
            elif "제외" in h:
                meta["excluded_flag"] = str(v)

        # Inject fiscal year
        if fiscal_year:
            meta["fiscal_year"] = fiscal_year

        companies.append({
            "company_name": name,
            "company_metadata": meta if meta else None,
            "rank_position": rank,
            "fill_color": fill_color,
        })
        rank += 1

    return companies


def find_header_row(ws, max_search=20):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_search, values_only=True), 1):
        non_none = [v for v in row if v is not None]
        if len(non_none) >= 4:
            text = " ".join(str(v) for v in non_none)
            if any(kw in text for kw in ["기업명", "순번", "회사명", "No", "NO"]):
                return i
    return None


def import_lattice(token, tasks):
    """래티스: F_V2 file, rows 1-29 only (before row 29 divider).
    Buyer excluded: 이안페이, 에버페이롤, 뉴플로이, 휴먼컨설팅그룹, 파워엠,
    시너지컨설팅, e-제너두, 메타페이, 아라인.
    Only 7 companies buyer was satisfied with."""
    fp = DRIVE_BASE + "1. 완료/래티스/아웃바운드서비스 기업 리스트_래티스_250825_F_V2.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["래티스"]
    header = find_header_row(ws)
    if not header:
        print("래티스: header not found")
        return

    # Extract up to row 29 (before divider), column A has "제외" markers
    excluded_names = {"이안페이", "에버페이롤", "뉴플로이", "휴먼컨설팅그룹", "파워엠",
                      "시너지컨설팅", "e-제너두", "메타페이", "아라인"}

    raw = extract_rows(ws, header, max_data_row=header + 29)

    # Check "제외" column (first column often has exclusion markers)
    results = []
    for c in raw:
        name = c["company_name"]
        # Skip if name contains excluded keywords
        if any(ex in name for ex in excluded_names):
            continue
        # Also check if excluded_flag is set
        meta = c.get("company_metadata") or {}
        if meta.get("excluded_flag") and "제외" in str(meta.get("excluded_flag", "")):
            continue
        results.append({
            "company_name": name,
            "company_metadata": c["company_metadata"],
            "rank_position": len(results) + 1,
        })

    task_id = tasks.get("A-03")
    if task_id:
        ok, err = upload_results(token, task_id, results)
        print(f"래티스 → A-03: {len(results)} companies {'✓' if ok else '✗ ' + err}")
        for r in results:
            print(f"  {r['rank_position']}. {r['company_name']}")
    return results


def import_andis(token, tasks):
    """앤디스파트너스: '리스트 합본 V2' sheet, grey-shaded rows = buyer rejected."""
    fp = DRIVE_BASE + "1. 완료/앤디스파트너스/아웃바운드서비스_앤디스파트너스_휴대폰&이메일.xlsx"
    wb = openpyxl.load_workbook(fp)
    ws = wb["리스트 합본 V2"]
    header = find_header_row(ws)
    if not header:
        print("앤디스: header not found")
        return

    raw = extract_rows(ws, header)

    # Filter out grey rows (rejected by buyer)
    results = []
    rejected = []
    for c in raw:
        fill = c.get("fill_color", "")
        is_grey = False
        if fill and fill not in ("00000000", "FFFFFFFF", "00FFFFFF"):
            if fill.startswith("FF") and len(fill) == 8:
                r = int(fill[2:4], 16)
                g = int(fill[4:6], 16)
                b = int(fill[6:8], 16)
                if abs(r - g) < 30 and abs(g - b) < 30 and r > 100:
                    is_grey = True
        if is_grey:
            rejected.append(c["company_name"])
        else:
            results.append({
                "company_name": c["company_name"],
                "company_metadata": c["company_metadata"],
                "rank_position": len(results) + 1,
            })

    task_id = tasks.get("A-04")
    if task_id:
        ok, err = upload_results(token, task_id, results)
        print(f"앤디스 → A-04: {len(results)} accepted, {len(rejected)} rejected {'✓' if ok else '✗ ' + err}")


def import_coemtech(token, tasks):
    """코엠테크: All companies satisfactory."""
    fp = DRIVE_BASE + "1. 완료/코엠테크/아웃바운드서비스 기업 리스트_코엠테크_ver1_251027.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = find_header_row(ws)
    if not header:
        print("코엠테크: header not found")
        return

    raw = extract_rows(ws, header)
    results = [{
        "company_name": c["company_name"],
        "company_metadata": c["company_metadata"],
        "rank_position": c["rank_position"],
    } for c in raw]

    task_id = tasks.get("A-05")
    if task_id:
        ok, err = upload_results(token, task_id, results)
        print(f"코엠테크 → A-05: {len(results)} companies {'✓' if ok else '✗ ' + err}")


def import_wefun1(token, tasks):
    """위펀1: 법정의무교육 롱리스트. 9 companies selected for tapping.
    Condition: 영업이익률 5-8%+"""
    fp = DRIVE_BASE + "위펀/위펀_법정의무교육 잠재 매수기업 롱리스트 명단.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = find_header_row(ws)
    if not header:
        print("위펀1: header not found")
        return

    raw = extract_rows(ws, header)

    # The 9 tapping targets specified by buyer
    tapping_targets = {
        "(주)휴넷", "(주)케이지에듀원", "(주)유비온", "(주)현대경제연구원",
        "(주)고려아카데미컨설팅", "(주)알파코", "(주)이패스코리아",
        "(주)유밥", "(주)한국이러닝교육센터",
    }

    results_all = [{
        "company_name": c["company_name"],
        "company_metadata": {
            **(c["company_metadata"] or {}),
            "tapping_target": c["company_name"] in tapping_targets or
                any(t in c["company_name"] for t in ["휴넷", "케이지에듀원", "유비온", "현대경제연구원",
                    "고려아카데미", "알파코", "이패스코리아", "유밥", "한국이러닝"]),
        },
        "rank_position": c["rank_position"],
    } for c in raw]

    # Use A-02 (위펀 query: B2B 교육/법정의무교육)
    task_id = tasks.get("A-02")
    if task_id:
        ok, err = upload_results(token, task_id, results_all)
        tapping_count = sum(1 for r in results_all if r["company_metadata"].get("tapping_target"))
        print(f"위펀1 → A-02: {len(results_all)} companies ({tapping_count} tapping targets) {'✓' if ok else '✗ ' + err}")


def import_wefun2(token, tasks):
    """위펀2: B2B 교육 + B2B 퀵, 10 each. 100% 경영권 인수."""
    fp = DRIVE_BASE + "위펀/아웃바운드서비스 기업 리스트_위펀_ver1_251231.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = find_header_row(ws)
    if not header:
        print("위펀2: header not found")
        return

    raw = extract_rows(ws, header)
    results = [{
        "company_name": c["company_name"],
        "company_metadata": c["company_metadata"],
        "rank_position": c["rank_position"],
    } for c in raw]

    # This maps to A-02 as well, but it's a separate list
    # Create a separate entry - use the wefun B2B quick query if it exists
    # For now attach to A-02 (위펀)
    # Actually A-02 already has wefun1 data, so let's skip duplication
    # We'll note this as supplementary data
    print(f"위펀2: {len(results)} companies (supplementary to A-02, not imported to avoid duplicate)")
    for r in results[:5]:
        cat = (r.get("company_metadata") or {}).get("category", "?")
        print(f"  {r['rank_position']}. [{cat}] {r['company_name']}")
    if len(results) > 5:
        print(f"  ... +{len(results)-5} more")


def import_payheer(token, tasks):
    """페이히어: TRS/렌탈 priority. Rows 1-19 are tapping targets."""
    fp = DRIVE_BASE + "페이히어/아웃바운드서비스 기업 리스트_페이히어_ver1_20260227.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = find_header_row(ws)
    if not header:
        print("페이히어: header not found")
        return

    raw = extract_rows(ws, header, max_data_row=header + 55)

    # Rows 1-19 are tapping targets (higher quality)
    results = []
    for c in raw:
        is_tapping = c["rank_position"] <= 19
        results.append({
            "company_name": c["company_name"],
            "company_metadata": {
                **(c["company_metadata"] or {}),
                "tapping_target": is_tapping,
            },
            "rank_position": c["rank_position"],
        })

    task_id = tasks.get("A-01")
    if task_id:
        ok, err = upload_results(token, task_id, results)
        tapping = sum(1 for r in results if r["company_metadata"].get("tapping_target"))
        print(f"페이히어 → A-01: {len(results)} companies ({tapping} tapping targets) {'✓' if ok else '✗ ' + err}")


def import_uilgijeon(token, tasks):
    """우일기전: 매도자 탐색."""
    fp = DRIVE_BASE + "1. 완료/우일기전/아웃바운드서비스 기업 리스트_우일기전_ver2_250930.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = find_header_row(ws)
    if not header:
        print("우일기전: header not found")
        return

    raw = extract_rows(ws, header)
    results = [{
        "company_name": c["company_name"],
        "company_metadata": c["company_metadata"],
        "rank_position": c["rank_position"],
    } for c in raw]

    task_id = tasks.get("D-02")
    if task_id:
        ok, err = upload_results(token, task_id, results)
        print(f"우일기전 → D-02: {len(results)} companies {'✓' if ok else '✗ ' + err}")


def import_utdam(token, tasks):
    """웃담에프엔비: 매도자 탐색."""
    fp = DRIVE_BASE + "1. 완료/웃담에프엔비/아웃바운드서비스 기업 리스트_웃담에프엔비_ver1_250925.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = find_header_row(ws)
    if not header:
        print("웃담에프엔비: header not found")
        return

    raw = extract_rows(ws, header)
    results = [{
        "company_name": c["company_name"],
        "company_metadata": c["company_metadata"],
        "rank_position": c["rank_position"],
    } for c in raw]

    task_id = tasks.get("D-03")
    if task_id:
        ok, err = upload_results(token, task_id, results)
        print(f"웃담에프엔비 → D-03: {len(results)} companies {'✓' if ok else '✗ ' + err}")


def main():
    token = login()
    tasks = get_all_tasks(token)
    print(f"Loaded {len(tasks)} tasks\n")

    import_lattice(token, tasks)
    print()
    import_andis(token, tasks)
    print()
    import_coemtech(token, tasks)
    print()
    import_wefun1(token, tasks)
    print()
    import_wefun2(token, tasks)
    print()
    import_payheer(token, tasks)
    print()
    import_uilgijeon(token, tasks)
    print()
    import_utdam(token, tasks)


if __name__ == "__main__":
    main()
