"""
Import golden list Excel files from Google Drive as labeling task results.

These are real query-answer pairs from completed CookieDeal outbound service customers.
Each customer folder contains an Excel with companies that were manually curated
as matches for that customer's search criteria.

Usage:
    python etl/scripts/import_golden_list.py [--dry-run] [--api-url http://localhost:8000]
"""

import argparse
import json
import os
import sys

import openpyxl
import requests

GOLDEN_LIST_BASE = (
    "/Users/seungohnam/Library/CloudStorage/GoogleDrive-nam.s@cookiedeal.io/"
    "Shared drives/Cookiedeal/Ops/아웃바운드 서비스/진행 고객/1. 완료/"
)

# Map customer folder names to query IDs in finetuning_queries_final.json
CUSTOMER_QUERY_MAP = {
    "래티스": "A-03",           # 급여 대행 업체, 매출 10억+, 수도권
    "코엠테크": "A-05",         # 정수기 솔레노이드밸브/온도·수위센서
    "앤디스파트너스": "A-04",    # 반도체 장비, 지분매각가 1,200억 이내
    "웃담에프엔비": "D-03",     # 웃담에프엔비 매도 — F&B 인수 의향
    "우일기전": "D-02",         # 우일기전 매도 — 기전 분야 시너지
    # These customers don't have direct query matches in the seed queries,
    # but we can create tasks for them
    "서울지공": None,
    "비저너리": None,
    "브이투": None,
    "이그나이트": None,
    "지앤엠컴퍼니(리스트 확정, 2025 10 17~)": None,
    "커넥토리얼": None,
}


def find_best_xlsx(folder_path: str) -> str | None:
    """Find the best (latest/final) xlsx file in a folder."""
    xlsx_files = [
        f for f in os.listdir(folder_path)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ]
    if not xlsx_files:
        return None

    # Prefer 확정 or 컨펌 files
    for f in xlsx_files:
        if "확정" in f or "컨펌" in f:
            return os.path.join(folder_path, f)

    # Otherwise pick latest by name
    xlsx_files.sort()
    return os.path.join(folder_path, xlsx_files[-1])


def extract_companies(filepath: str) -> list[dict]:
    """Extract company data from an outbound service Excel file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row
    header_row_idx = None
    headers = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        values = [cell.value for cell in row]
        non_none = [v for v in values if v is not None]
        if len(non_none) >= 4:
            # Check if this looks like a header
            text = " ".join(str(v) for v in non_none)
            if any(kw in text for kw in ["기업명", "순번", "회사명", "대표자", "사업자"]):
                header_row_idx = i
                headers = [str(v).strip().replace("\n", " ") if v else f"col_{j}" for j, v in enumerate(values)]
                break

    if not header_row_idx:
        print(f"  WARNING: Could not find header row in {filepath}")
        return []

    # Find key column indices
    name_col = None
    for j, h in enumerate(headers):
        if "기업명" in h or "회사명" in h:
            name_col = j
            break

    if name_col is None:
        # Try second column (often the company name)
        name_col = 2  # fallback

    companies = []
    rank = 1
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        row_list = list(row)
        if not any(v is not None for v in row_list[:6]):
            continue

        company_name = row_list[name_col] if name_col < len(row_list) else None
        if not company_name or not isinstance(company_name, str):
            continue

        company_name = company_name.strip()
        if not company_name:
            continue

        # Extract metadata from known column patterns
        metadata = {}
        for j, h in enumerate(headers):
            if j >= len(row_list) or row_list[j] is None:
                continue
            val = row_list[j]
            h_lower = h.lower()
            if "대표자" in h or "대표" in h:
                metadata["ceo"] = str(val)
            elif "사업자등록" in h:
                metadata["brn"] = str(val)
            elif "매출" in h:
                try:
                    metadata["revenue"] = int(float(val))
                except (ValueError, TypeError):
                    pass
            elif "영업이익" in h and "율" not in h:
                try:
                    metadata["operating_profit"] = int(float(val))
                except (ValueError, TypeError):
                    pass
            elif "자산" in h:
                try:
                    metadata["total_assets"] = int(float(val))
                except (ValueError, TypeError):
                    pass
            elif "기업유형" in h or "유형" in h:
                metadata["company_type"] = str(val)
            elif "홈페이지" in h:
                metadata["homepage"] = str(val)

        companies.append({
            "company_name": company_name,
            "company_metadata": metadata if metadata else None,
            "rank_position": rank,
        })
        rank += 1

    return companies


def main():
    parser = argparse.ArgumentParser(description="Import golden list as labeling task results")
    parser.add_argument("--dry-run", action="store_true", help="Print results without importing")
    parser.add_argument("--api-url", default="http://localhost:8000", help="Backend API URL")
    parser.add_argument("--email", default="admin@cookiedeal.com")
    parser.add_argument("--password", default="test1234")
    args = parser.parse_args()

    if not os.path.exists(GOLDEN_LIST_BASE):
        print(f"ERROR: Golden list base directory not found: {GOLDEN_LIST_BASE}")
        sys.exit(1)

    # Load query mapping from finetuning_queries_final.json
    queries_path = os.path.join(
        os.path.dirname(__file__), "..", "data", "finetuning_queries_final.json"
    )
    with open(queries_path) as f:
        queries_data = json.load(f)
    query_by_id = {q["id"]: q for q in queries_data["queries"]}

    # Login
    if not args.dry_run:
        login_resp = requests.post(
            f"{args.api_url}/api/labeling/auth/login",
            json={"email": args.email, "password": args.password},
        )
        login_resp.raise_for_status()
        token = login_resp.json()["access_token"]
        auth_headers = {"Authorization": f"Bearer {token}"}

        # Get task list to map query_id → task_id
        all_tasks = {}
        page = 1
        while True:
            resp = requests.get(
                f"{args.api_url}/api/labeling/tasks",
                params={"page": page, "size": 100},
                headers=auth_headers,
            )
            resp.raise_for_status()
            data = resp.json()
            for t in data["items"]:
                all_tasks[t["query_id"]] = t["id"]
            if page >= data["pages"]:
                break
            page += 1
        print(f"Loaded {len(all_tasks)} tasks from API")

    # Process each customer folder
    total_imported = 0
    for folder_name, query_id in CUSTOMER_QUERY_MAP.items():
        folder_path = os.path.join(GOLDEN_LIST_BASE, folder_name)
        if not os.path.exists(folder_path):
            print(f"SKIP: {folder_name} — folder not found")
            continue

        xlsx_path = find_best_xlsx(folder_path)
        if not xlsx_path:
            print(f"SKIP: {folder_name} — no xlsx found")
            continue

        companies = extract_companies(xlsx_path)
        if not companies:
            print(f"SKIP: {folder_name} — no companies extracted")
            continue

        if query_id and query_id in query_by_id:
            q = query_by_id[query_id]
            print(f"\n{folder_name} → {query_id}: \"{q['text'][:60]}...\"")
        else:
            print(f"\n{folder_name} → (no matching query)")

        print(f"  File: {os.path.basename(xlsx_path)}")
        print(f"  Companies: {len(companies)}")

        if args.dry_run:
            for c in companies[:5]:
                meta = c.get("company_metadata", {}) or {}
                rev = meta.get("revenue")
                rev_str = f" (매출 {rev/1e8:.0f}억)" if rev else ""
                print(f"    {c['rank_position']}. {c['company_name']}{rev_str}")
            if len(companies) > 5:
                print(f"    ... +{len(companies) - 5} more")
            continue

        # Import via API
        if query_id and query_id in all_tasks:
            task_id = all_tasks[query_id]
            resp = requests.post(
                f"{args.api_url}/api/labeling/admin/tasks/{task_id}/results",
                json={"results": companies},
                headers=auth_headers,
            )
            if resp.ok:
                print(f"  ✓ Imported {len(companies)} results to task {task_id} ({query_id})")
                total_imported += len(companies)
            else:
                print(f"  ✗ Error: {resp.status_code} {resp.text[:200]}")
        else:
            print(f"  ⚠ No task found for query_id={query_id}, skipping import")

    print(f"\n{'DRY RUN — ' if args.dry_run else ''}Total: {total_imported} company results imported")


if __name__ == "__main__":
    main()
